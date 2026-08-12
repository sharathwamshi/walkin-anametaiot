import io
import openpyxl
from openpyxl.styles import Font, PatternFill

CANDIDATE_HEADERS = [
    "name *", "email *", "phone *", "college", "branch", "cgpa", "passout_year", "resume_link"
]

QUESTION_HEADERS = [
    "question_text", "option_a", "option_b", "option_c", "option_d",
    "correct_option (A/B/C/D)", "marks", "negative_marks"
]


def _styled_header_sheet(headers, sheet_title, sample_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    header_fill = PatternFill(start_color="1450A3", end_color="1450A3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(18, len(h) + 4)
    if sample_rows:
        for r, row in enumerate(sample_rows, start=2):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def candidate_template_bytes():
    sample = [
        ("Ananya Rao", "ananya.rao@example.com", "9900011122", "KSIT Bangalore",
         "Computer Science", "8.2", "2026", ""),
    ]
    return _styled_header_sheet(CANDIDATE_HEADERS, "Candidates", sample)


def question_template_bytes():
    sample = [
        ("What is the time complexity of binary search?", "O(n)", "O(log n)", "O(n^2)", "O(1)",
         "B", 1, 0),
    ]
    return _styled_header_sheet(QUESTION_HEADERS, "Questions", sample)


def parse_candidate_excel(file_stream):
    from app.utils.form_validation import EMAIL_RE  # shared with the public registration form

    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    parsed, errors = [], []
    for idx, row in enumerate(rows, start=2):
        if row is None or all(v in (None, "") for v in row):
            continue
        try:
            name, email, phone, college, branch, cgpa, passout_year, resume_link = (
                list(row) + [None] * (8 - len(row))
            )[:8]
            name = str(name).strip() if name else None
            email = str(email).strip() if email else None
            phone = str(phone).strip() if phone else None

            # Only name, email, and phone are mandatory — every other column is
            # collected when present but never required.
            missing = [label for label, val in (("name", name), ("email", email), ("phone", phone)) if not val]
            if missing:
                errors.append(f"Row {idx}: {', '.join(missing)} {'is' if len(missing) == 1 else 'are'} required")
                continue
            if not EMAIL_RE.match(email):
                errors.append(f"Row {idx}: '{email}' is not a valid email address")
                continue

            parsed.append({
                "name": name,
                "email": email,
                "phone": phone,
                "college": str(college).strip() if college else None,
                "branch": str(branch).strip() if branch else None,
                "cgpa": str(cgpa).strip() if cgpa else None,
                "passout_year": str(passout_year).strip() if passout_year else None,
                "resume_link": str(resume_link).strip() if resume_link else None,
            })
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
    return parsed, errors


def parse_question_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    parsed, errors = [], []
    for idx, row in enumerate(rows, start=2):
        if row is None or all(v in (None, "") for v in row):
            continue
        try:
            (qtext, a, b, c, d, correct, marks, neg) = (list(row) + [None] * (8 - len(row)))[:8]
            if not qtext or not correct:
                errors.append(f"Row {idx}: question_text and correct_option are required")
                continue
            correct = str(correct).strip().upper()[:1]
            if correct not in ("A", "B", "C", "D"):
                errors.append(f"Row {idx}: correct_option must be A/B/C/D")
                continue
            parsed.append({
                "question_text": str(qtext).strip(),
                "option_a": str(a) if a is not None else None,
                "option_b": str(b) if b is not None else None,
                "option_c": str(c) if c is not None else None,
                "option_d": str(d) if d is not None else None,
                "correct_option": correct,
                "marks": float(marks) if marks not in (None, "") else 1.0,
                "negative_marks": float(neg) if neg not in (None, "") else 0.0,
            })
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
    return parsed, errors
